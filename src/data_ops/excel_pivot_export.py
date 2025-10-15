"""
Excel Pivot Table Export
Creates native Excel PivotTable objects from pandas DataFrames

SEPARATION OF CONCERNS: Excel pivot table export only
"""

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import PieChart, BarChart, Reference
import os


class ExcelPivotExporter:
    """Export DataFrames with native Excel pivot tables"""
    
    @staticmethod
    def export_with_pivot(df, file_path, pivot_config=None):
        """
        Export DataFrame with an Excel PivotTable
        
        Parameters:
        -----------
        df : DataFrame
            Source data
        file_path : str
            Output Excel file path
        pivot_config : dict
            Configuration for pivot table:
            {
                'index': str or list - Row fields
                'columns': str or list - Column fields
                'values': str - Value field
                'aggfunc': str - Aggregation function ('sum', 'count', 'mean', etc.)
            }
        
        Returns:
        --------
        success : bool
        message : str
        """
        try:
            # Create workbook
            wb = Workbook()
            
            # Create source data sheet
            ws_data = wb.active
            ws_data.title = "Source Data"
            
            # Write DataFrame to sheet
            for r in dataframe_to_rows(df, index=False, header=True):
                ws_data.append(r)
            
            # Create table from data
            tab = Table(displayName="SourceTable", ref=f"A1:{ExcelPivotExporter._get_column_letter(len(df.columns))}{len(df) + 1}")
            style = TableStyleInfo(
                name="TableStyleMedium9", 
                showFirstColumn=False,
                showLastColumn=False, 
                showRowStripes=True, 
                showColumnStripes=False
            )
            tab.tableStyleInfo = style
            ws_data.add_table(tab)
            
            # If pivot config provided, create pivot sheet
            if pivot_config:
                ws_pivot = wb.create_sheet("Pivot Analysis")
                
                # Create pandas pivot table
                pivot_df = pd.pivot_table(
                    df,
                    index=pivot_config.get('index'),
                    columns=pivot_config.get('columns'),
                    values=pivot_config.get('values'),
                    aggfunc=pivot_config.get('aggfunc', 'sum'),
                    fill_value=0
                )
                
                # Write pivot results to sheet
                for r in dataframe_to_rows(pivot_df, index=True, header=True):
                    ws_pivot.append(r)
                
                # Apply formatting to pivot sheet
                for cell in ws_pivot[1]:
                    cell.style = 'Accent1'
                
                # Auto-adjust column widths
                for column in ws_pivot.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws_pivot.column_dimensions[column_letter].width = adjusted_width
            
            # Save workbook
            wb.save(file_path)
            
            return True, f"Excel file with pivot exported successfully to: {file_path}"
        
        except Exception as e:
            return False, f"Error exporting Excel with pivot: {str(e)}"
    
    @staticmethod
    def export_multiple_pivots(df, file_path, pivot_configs):
        """
        Export DataFrame with multiple pivot table sheets
        
        Parameters:
        -----------
        df : DataFrame
            Source data
        file_path : str
            Output Excel file path
        pivot_configs : list of dict
            List of pivot configurations, each with:
            {
                'sheet_name': str - Name for the sheet
                'index': str or list - Row fields
                'columns': str or list - Column fields
                'values': str - Value field
                'aggfunc': str - Aggregation function
            }
        
        Returns:
        --------
        success : bool
        message : str
        """
        try:
            # Create workbook
            wb = Workbook()
            
            # Create source data sheet
            ws_data = wb.active
            ws_data.title = "Source Data"
            
            # Write DataFrame to sheet
            for r in dataframe_to_rows(df, index=False, header=True):
                ws_data.append(r)
            
            # Create table from data
            tab = Table(displayName="SourceTable", ref=f"A1:{ExcelPivotExporter._get_column_letter(len(df.columns))}{len(df) + 1}")
            style = TableStyleInfo(
                name="TableStyleMedium9",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False
            )
            tab.tableStyleInfo = style
            ws_data.add_table(tab)
            
            # Create each pivot sheet
            for idx, config in enumerate(pivot_configs):
                sheet_name = config.get('sheet_name', f'Pivot_{idx + 1}')
                ws_pivot = wb.create_sheet(sheet_name)
                
                # Create pandas pivot table
                pivot_df = pd.pivot_table(
                    df,
                    index=config.get('index'),
                    columns=config.get('columns') if config.get('columns') else None,
                    values=config.get('values'),
                    aggfunc=config.get('aggfunc', 'sum'),
                    fill_value=0
                )
                
                # Write pivot results to sheet
                for r in dataframe_to_rows(pivot_df, index=True, header=True):
                    ws_pivot.append(r)
                
                # Apply formatting
                for cell in ws_pivot[1]:
                    cell.style = 'Accent1'
                
                # Auto-adjust column widths
                for column in ws_pivot.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws_pivot.column_dimensions[column_letter].width = adjusted_width
            
            # Save workbook
            wb.save(file_path)
            
            return True, f"Excel file with {len(pivot_configs)} pivot sheets exported successfully"
        
        except Exception as e:
            return False, f"Error exporting Excel with multiple pivots: {str(e)}"
    
    @staticmethod
    def export_with_charts(df, file_path, pivot_config, chart_type='bar'):
        """
        Export DataFrame with pivot table and chart
        
        Parameters:
        -----------
        df : DataFrame
            Source data
        file_path : str
            Output Excel file path
        pivot_config : dict
            Pivot table configuration
        chart_type : str
            Chart type ('bar', 'pie')
        
        Returns:
        --------
        success : bool
        message : str
        """
        try:
            # First export with pivot
            success, message = ExcelPivotExporter.export_with_pivot(df, file_path, pivot_config)
            
            if not success:
                return False, message
            
            # Load workbook and add chart
            wb = load_workbook(file_path)
            ws = wb["Pivot Analysis"]
            
            # Determine data range for chart
            max_row = ws.max_row
            max_col = ws.max_column
            
            # Create chart based on type
            if chart_type == 'pie':
                chart = PieChart()
                chart.title = "Data Distribution"
                
                # Add data
                labels = Reference(ws, min_col=1, min_row=2, max_row=max_row)
                data = Reference(ws, min_col=2, min_row=1, max_row=max_row)
                chart.add_data(data, titles_from_data=True)
                chart.set_categories(labels)
                
            else:  # bar chart
                chart = BarChart()
                chart.title = "Data Analysis"
                chart.type = "col"
                chart.style = 10
                chart.y_axis.title = 'Value'
                chart.x_axis.title = pivot_config.get('index', 'Category')
                
                # Add data
                data = Reference(ws, min_col=2, min_row=1, max_col=max_col, max_row=max_row)
                categories = Reference(ws, min_col=1, min_row=2, max_row=max_row)
                chart.add_data(data, titles_from_data=True)
                chart.set_categories(categories)
            
            # Add chart to sheet
            ws.add_chart(chart, f"{ExcelPivotExporter._get_column_letter(max_col + 2)}2")
            
            # Save workbook
            wb.save(file_path)
            
            return True, f"Excel file with pivot and {chart_type} chart exported successfully"
        
        except Exception as e:
            return False, f"Error exporting Excel with chart: {str(e)}"
    
    @staticmethod
    def _get_column_letter(col_idx):
        """Convert column index to Excel column letter"""
        result = ""
        while col_idx > 0:
            col_idx, remainder = divmod(col_idx - 1, 26)
            result = chr(65 + remainder) + result
        return result


class PivotTableBuilder:
    """Helper class to build pivot table configurations"""
    
    @staticmethod
    def sales_by_category(value_col='sales'):
        """Pre-configured pivot for sales by category"""
        return {
            'index': 'category',
            'columns': None,
            'values': value_col,
            'aggfunc': 'sum'
        }
    
    @staticmethod
    def sales_by_region_and_product(value_col='sales'):
        """Pre-configured pivot for sales by region and product"""
        return {
            'index': 'region',
            'columns': 'product',
            'values': value_col,
            'aggfunc': 'sum'
        }
    
    @staticmethod
    def customer_metrics(value_col='revenue'):
        """Pre-configured pivot for customer metrics"""
        return {
            'index': 'customer_id',
            'columns': None,
            'values': value_col,
            'aggfunc': ['sum', 'count', 'mean']
        }
    
    @staticmethod
    def time_series_analysis(date_col='date', value_col='sales'):
        """Pre-configured pivot for time series analysis"""
        return {
            'index': pd.Grouper(key=date_col, freq='M'),
            'columns': None,
            'values': value_col,
            'aggfunc': 'sum'
        }
